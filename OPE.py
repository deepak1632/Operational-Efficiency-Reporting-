"""
generate_operational_efficiency_report.py

This script creates an Excel-based Operational Efficiency Report with:
- KPI Dashboard (Total Revenue, Total Orders, Average Order Value)
- Product-wise Revenue (Bar Chart)
- Monthly Revenue Trend (Line Chart)

Author: Deepak Jangra
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font

# Load sales data
df = pd.read_csv("sales_data.csv")  # Replace with your CSV file path
df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce')
df['UnitPrice'] = pd.to_numeric(df['UnitPrice'], errors='coerce')

# Calculate Revenue & Month
df['Revenue'] = df['Quantity'] * df['UnitPrice']
df['OrderDate'] = pd.to_datetime(df['OrderDate'], errors='coerce')
df['Month'] = df['OrderDate'].dt.to_period('M')

# Aggregations
sales_by_product = df.groupby('Product')['Revenue'].sum().reset_index()
sales_by_month = df.groupby('Month')['Revenue'].sum().reset_index()

# Create Excel workbook
file_path = "Operational_Efficiency_Report.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "KPI Dashboard"

# Summary KPIs
total_rev = df['Revenue'].sum()
total_orders = len(df)
avg_order_value = df['Revenue'].mean()

ws['A1'] = "Operational Efficiency Report"
ws['A1'].font = Font(bold=True, size=14)
ws.append([])
ws.append(["Total Revenue", total_rev])
ws.append(["Total Orders", total_orders])
ws.append(["Average Order Value", avg_order_value])
ws.append([])

# Sales by Product table
ws.append(["Sales by Product", "Revenue"])
for row in sales_by_product.itertuples(index=False):
    ws.append(row)

# Bar Chart: Product Revenue
chart = BarChart()
chart.title = "Revenue by Product"
data = Reference(ws, min_col=2, min_row=7, max_row=6+len(sales_by_product))
cats = Reference(ws, min_col=1, min_row=7, max_row=6+len(sales_by_product))
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.y_axis.title = 'Revenue'
chart.x_axis.title = 'Product'
ws.add_chart(chart, "D7")

# Monthly Trend sheet
ws2 = wb.create_sheet(title="Monthly Trend")
ws2.append(["Month", "Revenue"])
for row in sales_by_month.itertuples(index=False):
    ws2.append([str(row.Month), row.Revenue])

# Line Chart: Monthly Revenue
line_chart = LineChart()
line_chart.title = "Revenue by Month"
data = Reference(ws2, min_col=2, min_row=2, max_row=1+len(sales_by_month))
cats = Reference(ws2, min_col=1, min_row=2, max_row=1+len(sales_by_month))
line_chart.add_data(data, titles_from_data=False)
line_chart.set_categories(cats)
line_chart.y_axis.title = "Revenue"
line_chart.x_axis.title = "Month"
ws2.add_chart(line_chart, "D2")

# Save Excel file
wb.save(file_path)
print(f"Report saved as {file_path}")
